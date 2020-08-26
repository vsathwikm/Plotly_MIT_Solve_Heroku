# for file reading
import base64
import datetime
import io

# for creating the new total_score.xlsx
from create_total_score import create_total_score_excel

# for auth
import dash_auth
from flask import Flask 

# for downloading excel file
import dash_table_experiments as dte
from flask import send_file
import zipfile

# for turning excel into csv
import csv
import xlrd
import sys

# for writing to confirmed matches excel sheet
from xlwt import Workbook
from xlutils.copy import copy # not sure if needed

# for app
import os
import dash_table
import dash_core_components as dcc
import dash_html_components as html
import plotly.express as px
import pandas as pd
import dash
import dash_bootstrap_components as dbc
import plotly.graph_objects as go

# writing to excel files
import openpyxl


# for adding basic Auth 
VALID_USERNAME_PASSWORD_PAIRS = {
    'mit': 'solve2020'
}

external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css']
styles = {
    'pre': {
        'border': 'thin lightgrey solid',
        'overflowX': 'scroll'
    }
}

# the Flask app
app = Flask(__name__) 
app = dash.Dash(__name__, external_stylesheets=external_stylesheets)
server = app.server 
auth = dash_auth.BasicAuth(
    app,
    VALID_USERNAME_PASSWORD_PAIRS
)




# Determines how many matches have been created so writing to the excel
# file with new matches is a smooth process
df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
partners_list = df['MENTOR'].tolist()
# GLOBAL VARIABLE USED TO COUNT NUMBER OF MATCHES
COUNT_OF_MATCHES = len(partners_list)

# Getter for COUNT_OF_MATCHES
def get_count_of_matches():
    global COUNT_OF_MATCHES
    return COUNT_OF_MATCHES

# Increments COUNT_OF_MATCHES
def increment_count_of_matches():
    global COUNT_OF_MATCHES
    COUNT_OF_MATCHES += 1

# Initially the excel sheet 'total_score.xlsx' is used for the dashboard
# until new data is uploaded to the dashboard. The 'total_score.xlsx' sheet
# is hard coded and stored in the root directory
hardcoded_file_total_score = pd.ExcelFile('total_score.xlsx')
df_total_score = hardcoded_file_total_score.parse('Sheet1')

# list of solvers from hard coded 'total_score.xlsx'
Solvers = list(df_total_score.columns[1:])
# List of partners from hard coded 'total_score.xlsx'
Partners = list(df_total_score["Org_y"])

# Creates the initial horizonatl bar graph that is displayed on the dashboard
total_fig = px.bar(df_total_score.sort_values(Solvers[0], ascending=False)[:5], x=Solvers[0], 
y="Org_y", labels = {'Org_y':'MENTOR',Solvers[0]:'Total Score'})
total_fig.update_layout(yaxis={'categoryorder':'total ascending'})
# Format the bar graph
total_fig.update_layout(
    autosize=True,
    # width=700,
    height=500,
    margin=dict(
        l=50,
        r=50,
        b=100,
        t=100,
        pad=4
    )
)

# Getting initial, hardcoded Solver Table from dropdown bar
solver_needs_df = pd.read_csv("unused_files/excel_to_csv/solver_team_data.csv")
selected_solver_row_info = solver_needs_df[solver_needs_df['Org']==Solvers[0]].dropna(axis='columns')
selected_solver_row_info_list = list(solver_needs_df[solver_needs_df['Org']==Solvers[0]].dropna(axis='columns'))

# Creates a dictionary of all the Solver info to put in the selected_solver_table
selected_solver_row_list = []
for col in selected_solver_row_info:
    ind_row_dict = {}
    ind_row_dict["label"] = col
    ind_row_dict["value"] = selected_solver_row_info[col]
    selected_solver_row_list.append(ind_row_dict)

# Getting initial, hardcoded Partner Table - will be blank initially
partner_data_df = pd.read_csv("unused_files/excel_to_csv/partner_data.csv")
selected_partner_row_info = partner_data_df[partner_data_df['Org']==Partners[0]].dropna(axis='columns')
selected_partner_row_info_list = list(partner_data_df[partner_data_df['Org']==Partners[0]].dropna(axis='columns'))


# Creates a dictionary to put all Solvers as options of drop down menu
solver_list_dict = []
for solver in Solvers:
    ind_solver_dict = {}
    ind_solver_dict["label"]=solver
    ind_solver_dict["value"]=solver
    solver_list_dict.append(ind_solver_dict)

# Method that generates tables
# Used for the selected_solver_table and clicked_on_partner_table
def generate_table(dataframe, max_rows=200):    
    '''
    inputs:
    dataframe, padas df - dataframe to output to a table
    max_rows, int - max amount of rows to print. Set at 100 to 
    be high enough to deal with any dataframe used in this dashboard
    '''
    return html.Table([
        html.Thead(
            html.Tr([html.Th(col) for col in dataframe.columns])
        ),
        html.Tbody([
            html.Tr([
                html.Td(dataframe.iloc[i][col]) for col in dataframe.columns
            ]) for i in range(min(len(dataframe), max_rows))
        ])
        
    ], 
)


# Allows the input excel files to be turned into csv files which will be used
# to calculate the information required for pairings in create_total_score.py
def ExceltoCSV(excel_file, csv_file_base_path, csv_folder = "uploaded_excel_to_csv/"):
    '''
    inputs:
    Excel file, str - name of the original file
    csv_file_base_path, str - folder location where each file will be stored
    excel_folder, str - a folder with this name will be created in csv_base_file_path with sheet names
                        with excel file. If excel file has sheets then each sheet will be stored as
                        a seperte csv file

    '''    
    if not os.path.isdir(csv_file_base_path+csv_folder): 
        os.mkdir(csv_file_base_path+csv_folder)
    full_path = csv_file_base_path+csv_folder
    workbook = xlrd.open_workbook(excel_file)

    for sheet_name in workbook.sheet_names():
     #   print('processing - ' + sheet_name)
        worksheet = workbook.sheet_by_name(sheet_name)
        csv_file_full_path = full_path + sheet_name.lower().replace(" - ", "_").replace(" ","_") + '.csv'
        csvfile = open(csv_file_full_path, 'w',encoding='utf-8')
        writetocsv = csv.writer(csvfile, quoting = csv.QUOTE_ALL, )
        for rownum in range(worksheet.nrows):
            writetocsv.writerow(
    #                 list(x.encode('utf-8') if type(x) == type(u'') else x for x in worksheet.row_values(rownum)
                    worksheet.row_values(rownum)
                )
            
        csvfile.close()
     #( "{} has been saved at {}".format(sheet_name, csv_file_full_path))


# Method used to parse files that are uploaded through the upload button
def parse_contents(contents, filename, date):
    '''
    contents - contents of the file being uploaded
    filename - name of the file being uploaded
    date - time the file is uploaded

    This method will take in the excel file that is
    uploaded and will create csv files of each sheet
    in the directory 'uploaded_excel_to_csv' which is 
    in the root directory

    It also has to potential to print out sheets that
    are uploaded
    '''
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    try:
        if 'csv' in filename:
            # Assume that the user uploaded a CSV file
            return html.Div([
            'Please upload an excel sheets.'
        ])
        elif 'xls' in filename:
            # Assume that the user uploaded an excel file
            df = pd.read_excel(io.BytesIO(decoded))
    except Exception as e:
        print(e)
        return html.Div([
            'There was an error processing this file.'
        ])
    ExceltoCSV(excel_file=filename , csv_file_base_path ="" )
    return None


# APP LAYOUT
app.layout = html.Div(children=[
    html.H1(
        children='MIT SOLVE',
        style={
            'textAlign': 'center'
            
        }
    ),

    # Upload files button
    dcc.Upload(
        id='upload-data',
        children=html.Button('Upload Excel Data File'),
        style={
            'height': '60px',
            'textAlign': 'center',
        },
        # Allow multiple files to be uploaded
        multiple=True
    ),

    # Download all excel files button
    html.Div(children=[
        html.A(html.Button('Download All Excel Files'), href="/download_all/",
        ),
    ],
    style={
        'height': '60px',
        'textAlign': 'center',
    }),

    # Solver drop down menu 
    html.Label('Select a Solver'),
        dcc.Dropdown(
            id='Solver_dropdown',
            options= solver_list_dict,
            value = solver_list_dict[0]['value'], 
           ),

    # A few line breaks to make dashboard less crowded
    html.P(children=html.Br(), style={'textAlign': 'center'}),
    html.P(children=html.Br(), style={'textAlign': 'center'}),    

    # 2 side by side graphs
    html.Div([
        html.Div([
            # Title for the horizontal bar graph
             html.H2(children='Total Outputs Graph', style={'textAlign': 'center'}),
             # Horizontal total outputs graph
            dcc.Graph( 
                id='output_bargraph',
                figure= total_fig
            ),
        ], className="six columns"),

        html.Div([
            html.H3(id='individual_graph_title', children='Individual Graph', style={'textAlign': 'center'}),
            dcc.Graph(id='individual_graph', figure={'data': []})
        ], className="six columns"),
    ], className="row"),


    html.H4(id='weights_directions', children='Adjust Weight Values inside of Brackets Only --> [ ] '),
    # 2 side by side comment boxes for weights
    html.Div([
        html.Div([
            # Comment box 1
            dcc.Textarea(
                id='geo-weight',
                value='Textbox1', # initial value
                style={'display':'inline-block', 'width': '30%', 'height': '10%',},
            ),
        ], className="four columns"),
        html.Div([
            # Comment box 2
            dcc.Textarea(
                id='needs-weight',
                value='Textbox2', # initial value
                style={'display':'inline-block', 'width': '30%', 'height': '10%',},
            ),
        ], className="four columns")        
    ], className="row"),

    # 2 side by side comment boxes for weights
    html.Div([
        html.Div([
            # Comment box 3
            dcc.Textarea(
                id='challenges-weight',
                value='Textbox3', # initial value
                style={'display':'inline-block', 'width': '30%', 'height': '10%',},
            ),
        ], className="four columns"),
        html.Div([
            # Comment box 4
            dcc.Textarea(
                id='stage-weight',
                value='Textbox4', # initial value
                style={'display':'inline-block', 'width': '30%', 'height': '10%',},
            ),
        ], className="four columns"),      
    ], className="row"),
    html.P(children=html.Br(), style={'textAlign': 'center'}),
    html.Button('Submit Changes to Weights', id='submit-val', n_clicks=0),
    html.P(children=html.Br(), style={'textAlign': 'center'}),
    html.Div(id='confirmation-text',
             children='Press Submit to Edit Weights'),

    # Generates the table for the selected solver
    # selected_solver_row_info is that data of the seleced solver
    # that will go into the table
    html.H4(children='Selected Solver Information',style={'textAlign': 'center'}),
    dash_table.DataTable(
        id='selected_solver_table',
        columns=[{"name": i, "id": i} for i in selected_solver_row_info.columns],
        data=selected_solver_row_info.to_dict('records'),
        style_cell={
            'whiteSpace': 'normal',
            'height': 'auto',
            'textAlign': 'center',
            'font_family': 'helvetica',
            'font_size': '20px',
        },
        style_header={
        'backgroundColor': 'rgb(30, 30, 30)',
        'color': 'white',
        },
    ),

    # A few line breaks to make dashboard less crowded
    html.P(children=html.Br(), style={'textAlign': 'center'}),
    html.P(children=html.Br(), style={'textAlign': 'center'}),
   
    # Generate a checkbox that determines whether the current partner and solver are matched
    html.H3(children='Click Checkbox to Confirm Match between Selected Solver and Selected Partner',
        style={'textAlign': 'center'}
    ),
    dcc.RadioItems(
        id='checkbox_confirm',
        options=[  
            {'label': 'Yes Match?', 'value': 'Confirm'},
            {'label': 'No Match?', 'value': 'Denied'}
        ],
        style={
            'textAlign': 'center',
        },
        value='Denied', # set intitial value to 'denied' which means no match
        inputStyle={"margin-right": "20px"},
        labelStyle={'display': 'inline-block'}
    ),  

    # A line break to make dashboard less crowded
    html.P(children=html.Br(), style={'textAlign': 'center'}),

    # Generates table for the partner that is clicked on in the graph
    # selected_partner_row_info is that data of the seleced partner
    # that will go into the table - initially this table won't be populated
    html.H4(children='Clicked on Partner Information',style={'textAlign': 'center'}),
    dash_table.DataTable(
        id='clicked_on_partner_table',
        columns=[{"name": i, "id": i} for i in selected_partner_row_info.columns],
        data=selected_partner_row_info.to_dict('records'),
        style_cell={
            'whiteSpace': 'normal',
            'height': 'auto',
            'textAlign': 'center',
            'font_family': 'helvetica',
            'font_size': '20px',
        },
        style_header={
        'backgroundColor': 'rgb(30, 30, 30)',
        'color': 'white',
        },
    ),
    html.H4(children='Green = 0-1 matches, Blue = 2-3 matches, Red = 4 or more matches',style={'textAlign': 'center'}),

    # A few line breaks to make dashboard less crowded
    html.P(children=html.Br(), style={'textAlign': 'center'}),
    html.P(children=html.Br(), style={'textAlign': 'center'}),

    # Used to print out the newly calculated total score dataframe from
    # the uploaded files. Should only be used for debugging and is not set 
    # to be functional right now
    html.Div(id='output-data-upload'),

    # Print the solver matches for the selected partner below the partner table
    html.Div(id='partner-matches-list',
    children=[]),

    # Break line to space out the dashboard
    html.P(children=html.Br(), style={'textAlign': 'center'}),

    # hidden layout which is target of callbacks that don't update anything but
    # plotly dash requires outputs for all callbacks
    html.Div(id='hidden-div', 
    ),

    # Comment box
    dcc.Textarea(
        id='textarea-for-comments',
        value='Text area for comments', # initial value
        style={'width': '50%', 'height': 200, 'Align-items': 'center'},
    ),
])


# This callback prints the current list of solver matches for the current selected partner
# If there are no matches it default prints
@app.callback(
    dash.dependencies.Output('partner-matches-list', 'children'),
    [dash.dependencies.Input('checkbox_confirm', 'value')],
    [dash.dependencies.State('output_bargraph', 'clickData'),
    dash.dependencies.State('Solver_dropdown', 'value')],
)
def list_matches_for_a_partner(value, clickData, solver_name):
    '''
    param: value (str) - defines whether there is a match between partner and solver
    param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
    param: solver_name (str) - name of the selected solver from the dropdown menu
    return: a str containing the current matches for the selected mentor or a default response
    '''
    df = pd.read_excel('mit_solve_confirmed_matches.xlsx')
    partners_list = df['MENTOR'].tolist()
    solvers_list = df['SOLVER'].tolist()

    # If no partner is selected
    if clickData == None:
        # defualt response
        return 'You need to select a partner'

    matches_list = []

    # populate a list with all the solvers currently matched with the partner
    for i in range(len(solvers_list)):
        if partners_list[i] == str(clickData['points'][0]['label']):
            matches_list.append(solvers_list[i])

    # If match is just created add the selected solver (solver_name) to the list
    if value == 'Confirm':
        if solver_name not in matches_list:
            matches_list.append(solver_name)

    # If match is just deleted remove the selected solver (solver_name) from the list
    if value == 'Denied':
        if solver_name in matches_list:
            matches_list.remove(solver_name)

    if matches_list == []:
        return 'no current matches for this partner'
    else:
        return "List of current matches for " + str(clickData['points'][0]['label']) + ": \n" + str(matches_list)


# This method allows for you to download all of the generated excel files as a zip file
# Files are challenge_match.xlsx, geo_match.xlsx, needs_match.xlsx, stage_match.xlsx,
# total_score_from_upload.xlsx and mit_solve_confirmed_matches.xlsx
# TODO make sure the correct files are being uploaded - think wrong ones are right now
@app.server.route('/download_all/')
def download_all():
    zipf = zipfile.ZipFile('app/MIT_Solve_Excel_Files.zip','w', zipfile.ZIP_DEFLATED)
    for root,dirs, files in os.walk('MIT_SOLVE_downloadable_excel_files/'):
        for file in files:
            zipf.write('MIT_SOLVE_downloadable_excel_files/'+file)
    zipf.write('mit_solve_confirmed_matches.xlsx')
    zipf.close()
    return send_file('MIT_Solve_Excel_Files.zip',
            mimetype = 'zip',
            attachment_filename= 'MIT_Solve_Excel_Files.zip',
            as_attachment = True)


# This method will update the table displaying more information
# on the partner that is clicked on in the graph
@app.callback(
    [dash.dependencies.Output('clicked_on_partner_table', 'data'),
    dash.dependencies.Output('clicked_on_partner_table', 'style_cell')],
    [dash.dependencies.Input('output_bargraph', 'clickData'),
    dash.dependencies.Input('checkbox_confirm', 'value'),
    ])
def display_click_data(clickData, value):
    '''
    param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
    param: value (str) - defines whether there is a match between partner and solver
    return: data (dict) - data to be displayed in the partner table
    return: style_cell (dict) - information on how to style text color of mentor table
    '''
    # Check to make sure a partnere is selected
    if clickData != None:
        partner_name = clickData['points'][0]['label']
        partner_data_df = pd.read_csv("uploaded_excel_to_csv/partner_data.csv")
        selected_partner_row_info = partner_data_df[partner_data_df['Org']==partner_name].dropna(axis='columns')
        generate_table(selected_partner_row_info)
        df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
        partners_list = df['MENTOR'].tolist()

        
        # This loop counts how many matches there are for the specific partner
        partner_matches_count = 0
        for i in range(len(partners_list)):
            if partners_list[i] == partner_name:
                partner_matches_count += 1

        # Pick color for color_code based on number of matches
        # STILL A LITTLE BUGGY, DOESN'T UPDATE LIVE
        if partner_matches_count <= 1:
            color_code = 'green'
        elif partner_matches_count == 2 or partner_matches_count == 3:
            color_code = 'blue'
        else:
            color_code = 'red'

        new_style = {
            'whiteSpace': 'normal',
            'height': 'auto',
            'textAlign': 'center',
            'font_family': 'helvetica',
            'font_size': '20px',
            'color' : color_code
        }

        return [selected_partner_row_info.to_dict('records'), new_style]
    return [None, {
        'whiteSpace': 'normal',
        'height': 'auto',
        'textAlign': 'center',
        'font_family': 'helvetica',
        'font_size': '20px',
        }]


# This method will update the table displaying more information
# on the partner that is clicked on in the graph and also create the 
# additional individual graph
@app.callback(
    [dash.dependencies.Output('individual_graph', 'figure'),
    dash.dependencies.Output('individual_graph_title', 'children')],
    [dash.dependencies.Input('output_bargraph', 'clickData'),],
    [dash.dependencies.State('Solver_dropdown', 'value')]
    )
def update_individual_graph(clickData, solver_name):
    '''
    param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
    param: solver_name (str) - name of the selected solver from the dropdown menu
    return: figure (Plotly Express Bar Chart) - individual graph of category values
    return: children (str) - customized title for individual graph
    '''
    # Check to make sure a partnere is selected
    if clickData != None:
        # Must get value for partner compared to solver in: geo, needs, stage, challenge
        partner_name = clickData['points'][0]['label']

        geo_df = pd.read_csv("unused_files/excel_to_csv/geo_match.csv")
        geo_value = float(geo_df[geo_df["Partners\Solvers"]==partner_name].iloc[0][solver_name])
        
        needs_df = pd.read_csv("unused_files/excel_to_csv/needs_match.csv")
        needs_value = float(needs_df[needs_df["Partners\Solvers"]==partner_name].iloc[0][solver_name])

        stage_df = pd.read_csv("unused_files/excel_to_csv/stage_match.csv")
        stage_value = float(stage_df[stage_df["Partners\Solvers"]==partner_name].iloc[0][solver_name])

        challenge_df = pd.read_csv("unused_files/excel_to_csv/challenge_match.csv")
        challenge_value = float(challenge_df[challenge_df["Partners\Solvers"]==partner_name].iloc[0][solver_name])

        partner_values_dict = {'Labels': ['Challenges Score', 'Needs Score', 'Geo Score * Stage Score',
        'Geo Score', 'Stage Score'], 'Scores': [10*challenge_value, needs_value, 100*geo_value*stage_value,
        10*geo_value, 10*stage_value]}

        ind_fig = px.bar(partner_values_dict, x='Scores', y='Labels')
        return_string = "Individual Graph for '" + str(partner_name) + "'"
        
        return [ind_fig, return_string]
    
    figure={'data': []}
    return [figure, '']

# This callback fills the textboxes with weights of the current solver/partner pairing selected
@app.callback(
    [dash.dependencies.Output('geo-weight', 'value'),
    dash.dependencies.Output('needs-weight', 'value'),
    dash.dependencies.Output('challenges-weight', 'value'),
    dash.dependencies.Output('stage-weight', 'value'),],
    [dash.dependencies.Input('Solver_dropdown', 'value'),
    dash.dependencies.Input('output_bargraph', 'clickData')]
)
def fill_weight_text_boxes(solver_name, clickData):
    '''
    param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
    param: solver_name (str) - name of the selected solver from the dropdown menu
    return: geo-weight value (str) - the text that will be displayed in the geo weight textbox
    return: needs-weight value (str) - the text that will be displayed in the needs weight textbox
    return: challenges-weight value (str) - the text that will be displayed in the challenges weight textbox
    return: stage-weight value (str) - the text that will be displayed in the stage weight textbox
    '''
    df = pd.read_excel('solver_partner_weights.xlsx')
    if clickData != None:
        partner_name = clickData['points'][0]['label']
        weights = str(df[df["Org_y"]==partner_name].iloc[0][solver_name])

        # Splicing out the specific weights from the comment box
        # These are all indexes of commas within the weights string
        first_comma = weights.index(',')                      
        second_comma = weights[(first_comma + 1):].index(',') + first_comma + 1
        third_comma = weights[(second_comma + 1):].index(',') + second_comma + 1

        # Use the indexes of the commas to read in the weights
        geo_weight = weights[0:first_comma]
        needs_weight = weights[(first_comma+1):second_comma]
        challenges_weight = weights[(second_comma+1):third_comma]
        stage_weight = weights[(third_comma+1):]

        return ['weight for geo: [' + geo_weight + ']', 'weight for needs: [' + needs_weight + ']', 
        'weight for challenge: [' + challenges_weight + ']', 'weight for stage: [' + stage_weight + ']']
    return ['Select Partner Please', 'Select Partner Please', 'Select Partner Please', 'Select Partner Please']
    


# This callback edits weights in the excel sheet when submit button pressed
# TODO: currently the 'successfully edited...' message doesn't go away when a new 
# pairing is selected
@app.callback(
    [dash.dependencies.Output('confirmation-text', 'children'),],
    [dash.dependencies.Input('submit-val', 'n_clicks')],
    [dash.dependencies.State('geo-weight', 'value'),
    dash.dependencies.State('needs-weight', 'value'),
    dash.dependencies.State('challenges-weight', 'value'),
    dash.dependencies.State('stage-weight', 'value'),
    dash.dependencies.State('Solver_dropdown', 'value'),
    dash.dependencies.State('output_bargraph', 'clickData'),
    dash.dependencies.State('confirmation-text', 'children')
    ]
)
def edit_excel_sheet_with_new_weights(button_children, 
    new_geo_weight, new_needs_weight, new_challenges_weight,
    new_stage_weight, solver_name, clickData, 
    current_confirmation_text):
    '''
    param: button_children (int) - number of clicks on button -> just used to activate callback and not used in code
    param: new_geo_weight (str) - entire text within geo weight textbox
    param: new_needs_weight (str) - entire text within needs weight textbox
    param: new_challenges_weight (str) - entire text within challenges weight textbox
    param: new_stage_weight (str) - entire text within stage weight textbox
    param: solver_name (str) - name of the selected solver from the dropdown menu
    param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
    param: current_confirmation_text (str) - editing success message located underneath submit button on dashboard
    return: children (str) - customized edit success response message
    '''
    if clickData != None:
        print('edited')
        partner_name = clickData['points'][0]['label']

        file = 'solver_partner_weights.xlsx'
        wb = openpyxl.load_workbook(filename=file)
        ws = wb.get_sheet_by_name('Sheet1')

        # create new string to put in excel sheet
        new_weights = new_geo_weight[17:-1] + ',' + new_needs_weight[19:-1] + ',' + new_challenges_weight[23:-1] + ',' + new_stage_weight[19:-1]


        xls_file = pd.ExcelFile(file)
        df = xls_file.parse('Sheet1')
        # list of solvers from hard coded 'total_score.xlsx'
        Solvers = list(df.columns[1:])
        # List of partners from hard coded 'total_score.xlsx'
        Partners = list(df["Org_y"])

        # overwrite the old weights with new ones
        partner_row_num = Partners.index(partner_name) + 2
        solver_col_num = Solvers.index(solver_name) + 2

        ws.cell(row=partner_row_num, column=solver_col_num).value = new_weights
        # Save the workbook
        wb.save(file)
        return ['Weights edited for pairing of ' + partner_name + " and " + solver_name]
    return ['']


# Callback that either checks off or leaves blank the checkbox when a new solver or
# partneris selected
@app.callback(
    dash.dependencies.Output('checkbox_confirm', 'value'),
    [dash.dependencies.Input('Solver_dropdown', 'value'),
    dash.dependencies.Input('output_bargraph', 'clickData')]
)
def check_or_uncheck_checkbox(solver_name, clickData):
    '''
    param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
    param: solver_name (str) - name of the selected solver from the dropdown menu
    return: checkbox_confirm value (str) - status of mentor/solver pairing
    '''
    df = pd.read_excel('mit_solve_confirmed_matches.xlsx')
    partners_list = df['MENTOR'].tolist()
    solvers_list = df['SOLVER'].tolist()

    # check to make sure there is a partner selected
    if clickData == None:
            return 'You need to select a partner'

    # iterate through list of solvers to find currently selected solver (solver_name)
    for i in range(len(solvers_list)):
        if solvers_list[i] == solver_name:
            # if the solver name is found check if its partner is the currently selected partner
            if partners_list[i] == clickData['points'][0]['label']:
                # This is  a match 
                return 'Confirm'

    # If we get here this is not a match
    return 'Denied'


# Callback that adds and deletes matches to the 'mit_solve_confirmed_matches.xlsx'
@app.callback(
    dash.dependencies.Output('hidden-div', 'children'),
    [dash.dependencies.Input('checkbox_confirm', 'value')],
    [dash.dependencies.State('Solver_dropdown', 'value'),
    dash.dependencies.State('output_bargraph', 'clickData')]
    )
def add_confirmed_match(checkbox, solver_name, clickData):
    '''
    param: checkbox (str) - defines whether there is a match between partner and solver
    param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
    param: solver_name (str) - name of the selected solver from the dropdown menu
    return: children (str) - irrelavent output, will never be printed out and is used to 
    comply with needing an Output for every callback
    '''
    # Check if we are adding a match
    if checkbox == 'Confirm':
        if clickData == None:
            return 'You need to select a partner'
        else:
            df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
            partners_list = df['MENTOR'].tolist()
            solvers_list = df['SOLVER'].tolist()

            # checks if already a match
            for i in range(len(solvers_list)):
                if solvers_list[i] == solver_name:
                    if partners_list[i] == clickData['points'][0]['label']:
                        # This is already a match 
                        return None

            file = 'mit_solve_confirmed_matches.xlsx'
            wb = openpyxl.load_workbook(filename=file)
            ws = wb.get_sheet_by_name('Sheet1')

            # count number of matches 
            # start the count at 1 to account for this match not being added to the sheet yet
            matches_count_for_partner = 1
            for i in range(len(partners_list)):
                if partners_list[i] == clickData['points'][0]['label']:
                    matches_count_for_partner += 1
                    
            # insert the partner and solver names, as well as datetime and number of matches
            time_right_now = datetime.datetime.now()
            ws['A' + str(COUNT_OF_MATCHES + 2)] = str(clickData['points'][0]['label'])
            ws['B' + str(COUNT_OF_MATCHES + 2)] = str(solver_name)
            ws['C' + str(COUNT_OF_MATCHES + 2)] = str(time_right_now)
            ws['D' + str(COUNT_OF_MATCHES + 2)] = str(matches_count_for_partner)
            # Save the workbook
            wb.save(file)
            #increment amount of total matches
            increment_count_of_matches()
            return ''

    # Check if we are removing a match
    if checkbox == 'Denied':
        if clickData == None:
            return 'No partner selected'
        else:
            df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
            partners_list = df['MENTOR'].tolist()
            solvers_list = df['SOLVER'].tolist()

            # checks if already a match
            for i in range(len(solvers_list)):
                if solvers_list[i] == solver_name:
                    if partners_list[i] == clickData['points'][0]['label']:
                        # This match needs to be deleted

                        file = 'mit_solve_confirmed_matches.xlsx'
                        wb = openpyxl.load_workbook(filename=file)
                        # Select the right sheet
                        ws = wb.get_sheet_by_name('Sheet1')
                        # insert the partner and solver name, datetime, and number of matches
                        ws['A' + str(i + 2)] = str('')
                        ws['B' + str(i + 2)] = str('')
                        ws['C' + str(i + 2)] = str('')
                        ws['D' + str(i + 2)] = str('')
                        # Save the workbook
                        wb.save(file)
                        # NEED TO DECREMENT NUMBER OF MATCHES HERE
                        # LOGIC MAY NEED WORK TOO, NOT SURE IF JUST DECREMENTING IS THE RIGHT MOVE


# This method updates the table displaying more information on a solver
# everytime a new solver is selected from the dropdown
@app.callback(
    dash.dependencies.Output('selected_solver_table', 'data'),
    [dash.dependencies.Input('Solver_dropdown', 'value')])
def update_solver_table(value):
    '''
    param: solver_name (str) - name of the selected solver from the dropdown menu
    return: data (dict) - a dictionary containing data that will populate the solver table
    '''
    # Checks if new files have been uploaded yet instead of hard coded
    try:
        solver_needs_df = pd.read_csv("uploaded_excel_to_csv/solver_team_data.csv")
        selected_solver_row_info = solver_needs_df[solver_needs_df['Org']==value].dropna(axis='columns')
        generate_table(selected_solver_row_info)  
        return selected_solver_row_info.to_dict('records')
    # Uses the hard coded Solver info if no files have been uploaded yet
    except:
        solver_needs_df = pd.read_csv("unused_files/excel_to_csv/solver_team_data.csv")
        selected_solver_row_info = solver_needs_df[solver_needs_df['Org']==value].dropna(axis='columns')
        generate_table(selected_solver_row_info)  
        return selected_solver_row_info.to_dict('records')


# This method updates the graph when a new solver is selected from the dropdown
@app.callback(
    dash.dependencies.Output('output_bargraph', 'figure'),
    [dash.dependencies.Input('Solver_dropdown', 'value')])
def update_graph_from_solver_dropdown(value):
    '''
    param: solver_name (str) - name of the selected solver from the dropdown menu
    return: figure (Plotly Express Bar Chart) - the graph for total scores displayed on the dashboard
    '''
    # Checks if new files have been uploaded yet instead of hard coded
    try:
        xls_file_total_score = pd.ExcelFile('MIT_SOLVE_downloadable_excel_files/total_score_from_upload.xlsx')
        uploaded_df_total_score = xls_file_total_score.parse('Sheet1')
    # Uses the hard coded files if no files have been uploaded yet
    except:
        uploaded_df_total_score = df_total_score
    # Sort and crop top 5 values for new selected solver
    total_fig = px.bar(uploaded_df_total_score.sort_values(value, ascending=False)[:5], x=value, 
    y="Org_y", labels = {'Org_y':'MENTOR',value:'Total Score'})
    total_fig.update_layout(yaxis={'categoryorder':'total ascending'})
    return total_fig

# This method will create csv files for each sheet
# from the uploaded file. The uploaded file must be in the format of
# a singular excel file consisting of 2 sheets, which are the 
# partner_data and solver_team_data in that order
@app.callback(
    dash.dependencies.Output('output-data-upload', 'children'),
    [dash.dependencies.Input('upload-data', 'contents')],
    [dash.dependencies.State('upload-data', 'filename'),
    dash.dependencies.State('upload-data', 'last_modified')])
def update_output(list_of_contents, list_of_names, list_of_dates):
    '''
    param: contents - all of the files uploaded
    param: filename - name of the uploaded file
    param: last_modified - the date in which the file was last modified
    return: irrelavent output, will never be printed out and is used to 
    comply with needing an Output for every callback
    '''
    if list_of_contents is not None:
        # list_of_uploaded_files is fully available here
        children = [
            parse_contents(c, n, d) for c, n, d in
            zip(list_of_contents, list_of_names, list_of_dates)]
        new_total_score = create_total_score_excel()
        new_total_score.insert(0, "Partners", Partners, True)
        return None


# This callback will create a new bar chart with the data from the uploaded excel
# files instead of the preloaded old excel files
@app.callback(
    dash.dependencies.Output('Solver_dropdown', 'value'),
    [dash.dependencies.Input('upload-data', 'contents')],
)
def point_graph_to_uploaded_files(contents):
    '''
    param: contents - all of the files uploaded
    return: solver_name (str) - name of the selected solver from the dropdown menu
    '''
    try:
        # create new df from uploaded file
        xls_file_total_score = pd.ExcelFile('MIT_SOLVE_downloadable_excel_files/total_score_from_upload.xlsx')
        uploaded_df_total_score = xls_file_total_score.parse('Sheet1')
        # Create new graph with uploaded data instead of hardcoded
        new_solvers = list(uploaded_df_total_score.columns[1:])
        # Returns a different solver its obvious whether this worked or not
        return new_solvers[8]
    except:
        return Solvers[0]


# adding external stylesheets for 2 side by side graphs
app.css.append_css({
    'external_url': 'https://codepen.io/chriddyp/pen/bWLwgP.css'
})

if __name__ == '__main__':
    app.run_server(debug=True)