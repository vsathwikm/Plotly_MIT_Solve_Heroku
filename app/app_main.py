# for file reading
import base64
import datetime
import io

# for creating the new total_score.xlsx
from create_total_score import create_total_score_excel

# for auth
import dash_auth
from flask import Flask 

# for downloading excel file
from flask import send_file
import zipfile

# for turning excel into csv
import csv
import xlrd
import sys

import shutil

# for app
import os
import dash_table
import dash_core_components as dcc
import dash_html_components as html
import plotly.express as px
import pandas as pd
import dash
import plotly.graph_objects as go
from dash.dependencies import Output, Input
# writing to excel files
import openpyxl
import yaml 
from dash.exceptions import PreventUpdate
import callbacks 

with open("config.yml") as config_file: 
     config = yaml.load(config_file, Loader=yaml.FullLoader)





# Determines how many matches have been created so writing to the excel
# file with new matches is a smooth process
df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
partners_list = df['PARTNER'].tolist()
# GLOBAL VARIABLE USED TO COUNT NUMBER OF MATCHES
COUNT_OF_MATCHES = len(partners_list)

# Getter for COUNT_OF_MATCHES
def get_count_of_matches():
    global COUNT_OF_MATCHES
    return COUNT_OF_MATCHES

# Increments COUNT_OF_MATCHES
def increment_count_of_matches():
    global COUNT_OF_MATCHES
    COUNT_OF_MATCHES += 1

# Initially the excel sheet 'total_score.xlsx' is used for the dashboard
# until new data is uploaded to the dashboard. The 'total_score.xlsx' sheet
# is hard coded and stored in the root directory
# hardcoded_file_total_score = pd.ExcelFile('../total_score.xlsx')
# df_total_score = hardcoded_file_total_score.parse('Sheet1')


selected_solver_row_list = []
# Creates a dictionary to put all Solvers as options of drop down menu
solver_list_dict = []

# Method that generates tables
# Used for the selected_solver_table and clicked_on_partner_table

# Allows the input excel files to be turned into csv files which will be used
# to calculate the information required for pairings in create_total_score.py





################################ START APP ################################################
###########################################################################################



################################ END APP ################################################
#########################################################################################




# @app.callback([dash.dependencies.Output('checkbox_confirm', 'options'),
#               dash.dependencies.Output('checkbox_confirm', 'value')], 
#             [dash.dependencies.Input('solver-dropdown', 'value')])
# def generate_checkboxes(solver): 
 
#     # Checks if new files have been uploaded yet instead of hard coded
#     xls_file_total_score = pd.read_excel(config['total_score_location'])
#     uploaded_df_total_score = xls_file_total_score.copy()
#     filtered_data = uploaded_df_total_score.sort_values(solver, ascending=False)[:config['max_matches']]
     
#     checkbox = [ {'label': "Yes", 'value': "Yes"}, 
#                  {'label': "No", 'value': "No"}] 
#     value = "No"
    
#     return [checkbox, value]

# @app.callback(
#     dash.dependencies.Output('clicked_on_partner_table', 'style_cell'),
#     [dash.dependencies.Input('checkbox_confirm', 'value'),
#      dash.dependencies.Input('solver-dropdown', 'value'), 
#     dash.dependencies.Input('output_bargraph', 'clickData')])
# def update_confirmed(selected_partners, solver, table_partner): 
    
#     style_cell={
#             'whiteSpace': 'normal',
#             'height': 'auto',
#             'textAlign': 'center',
#             'font_family': 'helvetica',
#             'font_size': '20px',
#         }    
#     partner_name =  table_partner['points'][0]['y']

#     # Make partners list if it does not exist 
#     if not os.path.exists(config['track_partners']): 
#         partners_list = pd.read_excel(config['total_score_location'])['Org_y'].values.tolist()
       
#         solvers =[ '' for x in range(0,len(partners_list))]
#         count = [0 for x in range(0,len(partners_list))]
#         partners_trackers = pd.DataFrame(data=[partners_list, solvers, count],
#                                          index=['partners','solvers', 'counter']).T
#         partners_trackers.to_csv(config['track_partners'], index=False)
#     else: 
#         partners_trackers = pd.read_csv(config['track_partners'])
        
#     # check if confirmed matches file exists, if not then create it
#     # while creating, label partner-solver matches
#     # also make additions to the partners list
#     if not os.path.exists(config['confirmed_matches']): 
#         total_score = pd.read_excel(config['total_score_location'])
#         for col in total_score:
#             if not col == 'Org_y':
#                 total_score[col].values[:] = 0 

#         if selected_partners == "Yes": 

#             total_score[solver][total_score['Org_y'] == partner_name] = 1
#             partners_trackers['counter'][partners_trackers['partners'] == partner_name] += 1
#             print(partners_trackers['solvers'][partners_trackers['partners'] == partner_name])
#             partners_trackers['solvers'][partners_trackers['partners']==partner_name] += ', '+solver  
        
#         total_score.to_csv(config['confirmed_matches'], index=False)
#         partners_trackers.to_csv(config['track_partners'], index=False)
#         solvers_for_partner = int(total_score.loc[total_score['Org_y']==partner_name].sum(axis=1).values) 
#         if solvers_for_partner <= config['partner_inter'] : 
#             style_cell['color'] = 'green'
#         elif solvers_for_partner > config['partner_inter']  and solvers_for_partner <= config['max_matches']: 
#             style_cell['color'] = 'blue'
#         else: 
#             style_cell['color'] = 'red'

#         return style_cell

#     else: 
#         total_score = pd.read_csv(config['confirmed_matches'])
        
#         if selected_partners == "Yes": 
#             total_score[solver][total_score['Org_y']== partner_name] = 1
#             partners_trackers['counter'][partners_trackers['partners'] == partner_name] += 1
#             print(partners_trackers['solvers'][partners_trackers['partners'] == partner_name])
#             partners_trackers['solvers'][partners_trackers['partners']==partner_name] += ', '+solver  

#         solvers_for_partner = int(total_score.loc[total_score['Org_y']==partner_name].sum(axis=1).values) 
#         partners_trackers.to_csv(config['track_partners'], index=False)
#         total_score.to_csv(config['confirmed_matches'], index=False)
        
#         if solvers_for_partner <= config['partner_inter'] : 
#             style_cell['color'] = 'green'
#         elif solvers_for_partner > config['partner_inter']  and solvers_for_partner <= config['max_matches']: 
#             style_cell['color'] = 'blue'
#         else: 
#             style_cell['color'] = 'red'

#         return style_cell



@app.callback(
    [dash.dependencies.Output('geo-weight', 'value'),
    dash.dependencies.Output('needs-weight', 'value'),
    dash.dependencies.Output('challenges-weight', 'value'),
    dash.dependencies.Output('stage-weight', 'value'),],
    [dash.dependencies.Input('solver-dropdown', 'value'),
    dash.dependencies.Input('output_bargraph', 'clickData')]
)
def fill_weight_text_boxes(solver_name, clickData):
    '''
    param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
    param: solver_name (str) - name of the selected solver from the dropdown menu
    return: geo-weight value (str) - the text that will be displayed in the geo weight textbox
    return: needs-weight value (str) - the text that will be displayed in the needs weight textbox
    return: challenges-weight value (str) - the text that will be displayed in the challenges weight textbox
    return: stage-weight value (str) - the text that will be displayed in the stage weight textbox
    '''

    if not os.path.exists(config['weights']): 
        weights_matrix = pd.read_excel(config['total_score_location'])
        for col in weights_matrix:
            if not col == 'Org_y':
                weights_matrix[col].values[:] = {'label':0, 'ere':3}
        weights_matrix.to_excel(config['weights'])
    
    df = pd.read_excel(config['weights'])
    if clickData != None:
        partner_name = clickData['points'][0]['label']
        weights = str(df[df["Org_y"]==partner_name].iloc[0][solver_name])
        # Splicing out the specific weights from the comment box
        # These are all indexes of commas within the weights string
        # first_comma = weights.index(',')                      
        # second_comma = weights[(first_comma + 1):].index(',') + first_comma + 1
        # third_comma = weights[(second_comma + 1):].index(',') + second_comma + 1

        # Use the indexes of the commas to read in the weights
        # geo_weight = weights[0:first_comma]
        # needs_weight = weights[(first_comma+1):second_comma]
        # challenges_weight = weights[(second_comma+1):third_comma]
        # stage_weight = weights[(third_comma+1):]
        if str(weights) == "0": 
            weights_list = ["0","0","0","0"]
        else:
            weights_list = weights.split()
        
        geo_weight = weights_list[0]
        needs_weight = weights_list[1]
        challenges_weight = weights_list[2]
        stage_weight = weights_list[3]

        return [geo_weight, needs_weight, challenges_weight, stage_weight]
    return ['Select Partner Please', 'Select Partner Please', 'Select Partner Please', 'Select Partner Please']
    

@app.callback(
    [dash.dependencies.Output('confirmation-text', 'children'),],
    [dash.dependencies.Input('submit-val', 'n_clicks')],
    [dash.dependencies.State('geo-weight', 'value'),
    dash.dependencies.State('needs-weight', 'value'),
    dash.dependencies.State('challenges-weight', 'value'),
    dash.dependencies.State('stage-weight', 'value'),
    dash.dependencies.State('solver-dropdown', 'value'),
    dash.dependencies.State('output_bargraph', 'clickData'),
    dash.dependencies.State('confirmation-text', 'children')
    ]
)    
def edit_excel_sheet_with_new_weights(button_children, 
    new_geo_weight, new_needs_weight, new_challenges_weight,
    new_stage_weight, solver_name, clickData, 
    current_confirmation_text):
    '''
    param: button_children (int) - number of clicks on button -> just used to activate callback and not used in code
    param: new_geo_weight (str) - entire text within geo weight textbox
    param: new_needs_weight (str) - entire text within needs weight textbox
    param: new_challenges_weight (str) - entire text within challenges weight textbox
    param: new_stage_weight (str) - entire text within stage weight textbox
    param: solver_name (str) - name of the selected solver from the dropdown menu
    param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
    param: current_confirmation_text (str) - editing success message located underneath submit button on dashboard
    return: children (str) - customized edit success response message
    '''
 
    if clickData != None:
        print('edited')
        partner_name = clickData['points'][0]['y']

        file = config['weights']
        wb = pd.read_excel(file)
        # ws = wb.get_sheet_by_name('Sheet1')

        # create new string to put in excel sheet
        new_weights = str(new_geo_weight) + ',' + str(new_needs_weight) + ',' + str(new_challenges_weight)+ ',' + str(new_stage_weight)
        print("new weights",  new_weights)
        # print(wb[wb['Org_y'] == partner_name].loc[solver_name])
        print(wb[wb['Org_y'] == partner_name])
        print(wb[wb['Org_y'] == partner_name].loc[solver_name])
        print("done")
        wb[wb['Org_y'] == partner_name][solver_name] = {'label':0, 'ere':3}
        
        wb.to_excel(config['weights'], index=False)
        # xls_file = pd.ExcelFile(file)
        # df = xls_file.parse('Sheet1')
        # # list of solvers from hard coded 'total_score.xlsx'
        # Solvers = list(df.columns[1:])
        # # List of partners from hard coded 'total_score.xlsx'
        # Partners = list(df["Org_y"])


        # # overwrite the old weights with new ones
        # partner_row_num = Partners.index(partner_name) + 2
        # solver_col_num = Solvers.index(solver_name) + 2

        # ws.cell(row=partner_row_num, column=solver_col_num).value = new_weights
        # # Save the workbook
        # wb.save(file)

        return ['Weights edited for pairing of ' + partner_name + " and " + solver_name]
    return ['']



# # Callback that adds and deletes matches to the 'mit_solve_confirmed_matches.xlsx'
# @app.callback(
#     dash.dependencies.Output('hidden-div', 'children'),
#     [dash.dependencies.Input('checkbox_confirm', 'value')],
#     [dash.dependencies.State('solver-dropdown', 'value'),
#     dash.dependencies.State('output_bargraph', 'clickData')]
#     )
# def add_confirmed_match(checkbox, solver_name, clickData):
#     '''
#     param: checkbox (str) - defines whether there is a match between partner and solver
#     param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
#     param: solver_name (str) - name of the selected solver from the dropdown menu
#     return: children (str) - irrelavent output, will never be printed out and is used to 
#     comply with needing an Output for every callback
#     '''
#     # Check if we are adding a match
#     if checkbox == 'Confirm':
#         if clickData == None:
#             return 'You need to select a partner'
#         else:
#             df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
#             partners_list = df['PARTNER'].tolist()
#             solvers_list = df['SOLVER'].tolist()

#             # checks if already a match
#             for i in range(len(solvers_list)):
#                 if solvers_list[i] == solver_name:
#                     if partners_list[i] == clickData['points'][0]['label']:
#                         # This is already a match 
#                         return None

#             file = 'mit_solve_confirmed_matches.xlsx'
#             wb = openpyxl.load_workbook(filename=file)
#             ws = wb.get_sheet_by_name('Sheet1')

#             # count number of matches 
#             # start the count at 1 to account for this match not being added to the sheet yet
#             matches_count_for_partner = 1
#             for i in range(len(partners_list)):
#                 if partners_list[i] == clickData['points'][0]['label']:
#                     matches_count_for_partner += 1
                    
#             # insert the partner and solver names, as well as datetime and number of matches
#             time_right_now = datetime.datetime.now()
#             ws['A' + str(COUNT_OF_MATCHES + 2)] = str(clickData['points'][0]['label'])
#             ws['B' + str(COUNT_OF_MATCHES + 2)] = str(solver_name)
#             ws['C' + str(COUNT_OF_MATCHES + 2)] = str(time_right_now)
#             ws['D' + str(COUNT_OF_MATCHES + 2)] = str(matches_count_for_partner)
#             # Save the workbook
#             wb.save(file)
#             #increment amount of total matches
#             increment_count_of_matches()
#             return ''

#     # Check if we are removing a match
#     if checkbox == 'Denied':
#         if clickData == None:
#             return 'No partner selected'
#         else:
#             df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
#             partners_list = df['PARTNER'].tolist()
#             solvers_list = df['SOLVER'].tolist()

#             # checks if already a match
#             for i in range(len(solvers_list)):
#                 if solvers_list[i] == solver_name:
#                     if partners_list[i] == clickData['points'][0]['label']:
#                         # This match needs to be deleted

#                         file = 'mit_solve_confirmed_matches.xlsx'
#                         wb = openpyxl.load_workbook(filename=file)
#                         # Select the right sheet
#                         ws = wb.get_sheet_by_name('Sheet1')
#                         # insert the partner and solver name, datetime, and number of matches
#                         ws['A' + str(i + 2)] = str('')
#                         ws['B' + str(i + 2)] = str('')
#                         ws['C' + str(i + 2)] = str('')
#                         ws['D' + str(i + 2)] = str('')
#                         # Save the workbook
#                         wb.save(file)
#                         # NEED TO DECREMENT NUMBER OF MATCHES HERE
#                         # LOGIC MAY NEED WORK TOO, NOT SURE IF JUST DECREMENTING IS THE RIGHT MOVE






# # This callback prints the current list of solver matches for the current selected partner
# # If there are no matches it default prints
# @app.callback(
#     dash.dependencies.Output('partner-matches-list', 'children'),
#     [dash.dependencies.Input('checkbox_confirm', 'value')],
#     [dash.dependencies.State('output_bargraph', 'clickData'),
#     dash.dependencies.State('solver-dropdown', 'value')],
# )
# def list_matches_for_a_partner(value, clickData, solver_name):
#     '''
#     param: value (str) - defines whether there is a match between partner and solver
#     param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
#     param: solver_name (str) - name of the selected solver from the dropdown menu
#     return: a str containing the current matches for the selected mentor or a default response
#     '''
#     df = pd.read_excel('mit_solve_confirmed_matches.xlsx')
#     partners_list = df['PARTNER'].tolist()
#     solvers_list = df['SOLVER'].tolist()

#     # If no partner is selected
#     if clickData == None:
#         # defualt response
#         return 'You need to select a partner'

#     matches_list = []

#     # populate a list with all the solvers currently matched with the partner
#     for i in range(len(solvers_list)):
#         if partners_list[i] == str(clickData['points'][0]['y']):
#             matches_list.append(solvers_list[i])

#     # If match is just created add the selected solver (solver_name) to the list
#     if value == 'Confirm':
#         if solver_name not in matches_list:
#             matches_list.append(solver_name)

#     # If match is just deleted remove the selected solver (solver_name) from the list
#     if value == 'Denied':
#         if solver_name in matches_list:
#             matches_list.remove(solver_name)

#     if matches_list == []:
#         return 'no current matches for this partner'
#     else:
#         return "List of current matches for " + str(clickData['points'][0]['y']) + ": \n" + str(matches_list)




# # This method will update the table displaying more information
# # on the partner that is clicked on in the graph
# @app.callback(
#     [dash.dependencies.Output('clicked_on_partner_table', 'data'),
#     dash.dependencies.Output('clicked_on_partner_table', 'style_cell')],
#     [dash.dependencies.Input('output_bargraph', 'clickData'),
#     dash.dependencies.Input('checkbox_confirm', 'value'),
#     ])
# def display_click_data(clickData, value):
#     '''
#     param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
#     param: value (str) - defines whether there is a match between partner and solver
#     return: data (dict) - data to be displayed in the partner table
#     return: style_cell (dict) - information on how to style text color of mentor table
#     '''
#     # Check to make sure a partnere is selected
#     if clickData != None:
#         partner_name = clickData['points'][0]['label']
#         partner_data_df = pd.read_csv("../uploaded_excel_to_csv/partner_data.csv")
#         selected_partner_row_info = partner_data_df[partner_data_df['Org']==partner_name].dropna(axis='columns')
#         generate_table(selected_partner_row_info)
#         df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
#         partners_list = df['PARTNER'].tolist()

        
#         # This loop counts how many matches there are for the specific partner
#         partner_matches_count = 0
#         for i in range(len(partners_list)):
#             if partners_list[i] == partner_name:
#                 partner_matches_count += 1

#         # Pick color for color_code based on number of matches
#         # STILL A LITTLE BUGGY, DOESN'T UPDATE LIVE
#         if partner_matches_count <= 1:
#             color_code = 'green'
#         elif partner_matches_count == 2 or partner_matches_count == 3:
#             color_code = 'blue'
#         else:
#             color_code = 'red'

#         new_style = {
#             'whiteSpace': 'normal',
#             'height': 'auto',
#             'textAlign': 'center',
#             'font_family': 'helvetica',
#             'font_size': '20px',
#             'color' : color_code
#         }

#         return [selected_partner_row_info.to_dict('records'), new_style]
#     return [None, {
#         'whiteSpace': 'normal',
#         'height': 'auto',
#         'textAlign': 'center',
#         'font_family': 'helvetica',
#         'font_size': '20px',
#         }]


# # This method will update the table displaying more information
# # on the partner that is clicked on in the graph and also create the 
# # additional individual graph
# @app.callback(
#     [dash.dependencies.Output('individual_graph', 'figure'),
#     dash.dependencies.Output('individual_graph_title', 'children')],
#     [dash.dependencies.Input('output_bargraph', 'clickData'),],
#     [dash.dependencies.State('Solver_dropdown', 'value')]
#     )
# def update_individual_graph(clickData, solver_name):
#     '''
#     param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
#     param: solver_name (str) - name of the selected solver from the dropdown menu
#     return: figure (Plotly Express Bar Chart) - individual graph of category values
#     return: children (str) - customized title for individual graph
#     '''
#     # Check to make sure a partnere is selected
#     if clickData != None:
#         # Must get value for partner compared to solver in: geo, needs, stage, challenge
#         partner_name = clickData['points'][0]['label']

#         geo_df = pd.read_csv("../unused_files/excel_to_csv/geo_match.csv")
#         geo_value = float(geo_df[geo_df["Partners\Solvers"]==partner_name].iloc[0][solver_name])
        
#         needs_df = pd.read_csv("../unused_files/excel_to_csv/needs_match.csv")
#         needs_value = float(needs_df[needs_df["Partners\Solvers"]==partner_name].iloc[0][solver_name])

#         stage_df = pd.read_csv("../unused_files/excel_to_csv/stage_match.csv")
#         stage_value = float(stage_df[stage_df["Partners\Solvers"]==partner_name].iloc[0][solver_name])

#         challenge_df = pd.read_csv("../unused_files/excel_to_csv/challenge_match.csv")
#         challenge_value = float(challenge_df[challenge_df["Partners\Solvers"]==partner_name].iloc[0][solver_name])

#         partner_values_dict = {'Labels': ['Challenges Score', 'Needs Score', 'Geo Score * Stage Score',
#         'Geo Score', 'Stage Score'], 'Scores': [10*challenge_value, needs_value, 100*geo_value*stage_value,
#         10*geo_value, 10*stage_value]}

#         ind_fig = px.bar(partner_values_dict, x='Scores', y='Labels')
#         return_string = "Individual Graph for '" + str(partner_name) + "'"
        
#         return [ind_fig, return_string]
    
#     figure={'data': []}
#     return [figure, '']


# # This callback edits weights in the excel sheet when submit button pressed
# # TODO: currently the 'successfully edited...' message doesn't go away when a new 
# # pairing is selected
# @app.callback(
#     [dash.dependencies.Output('confirmation-text', 'children'),],
#     [dash.dependencies.Input('submit-val', 'n_clicks')],
#     [dash.dependencies.State('geo-weight', 'value'),
#     dash.dependencies.State('needs-weight', 'value'),
#     dash.dependencies.State('challenges-weight', 'value'),
#     dash.dependencies.State('stage-weight', 'value'),
#     dash.dependencies.State('Solver_dropdown', 'value'),
#     dash.dependencies.State('output_bargraph', 'clickData'),
#     dash.dependencies.State('confirmation-text', 'children')
#     ]
# )
# def edit_excel_sheet_with_new_weights(button_children, 
#     new_geo_weight, new_needs_weight, new_challenges_weight,
#     new_stage_weight, solver_name, clickData, 
#     current_confirmation_text):
#     '''
#     param: button_children (int) - number of clicks on button -> just used to activate callback and not used in code
#     param: new_geo_weight (str) - entire text within geo weight textbox
#     param: new_needs_weight (str) - entire text within needs weight textbox
#     param: new_challenges_weight (str) - entire text within challenges weight textbox
#     param: new_stage_weight (str) - entire text within stage weight textbox
#     param: solver_name (str) - name of the selected solver from the dropdown menu
#     param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
#     param: current_confirmation_text (str) - editing success message located underneath submit button on dashboard
#     return: children (str) - customized edit success response message
#     '''
#     if clickData != None:
#         print('edited')
#         partner_name = clickData['points'][0]['label']

#         file = '../solver_partner_weights.xlsx'
#         wb = openpyxl.load_workbook(filename=file)
#         ws = wb.get_sheet_by_name('Sheet1')

#         # create new string to put in excel sheet
#         new_weights = new_geo_weight[17:-1] + ',' + new_needs_weight[19:-1] + ',' + new_challenges_weight[23:-1] + ',' + new_stage_weight[19:-1]


#         xls_file = pd.ExcelFile(file)
#         df = xls_file.parse('Sheet1')
#         # list of solvers from hard coded 'total_score.xlsx'
#         Solvers = list(df.columns[1:])
#         # List of partners from hard coded 'total_score.xlsx'
#         Partners = list(df["Org_y"])

#         # overwrite the old weights with new ones
#         partner_row_num = Partners.index(partner_name) + 2
#         solver_col_num = Solvers.index(solver_name) + 2

#         ws.cell(row=partner_row_num, column=solver_col_num).value = new_weights
#         # Save the workbook
#         wb.save(file)
#         return ['Weights edited for pairing of ' + partner_name + " and " + solver_name]
#     return ['']


# # Callback that either checks off or leaves blank the checkbox when a new solver or
# # partneris selected
# @app.callback(
#     dash.dependencies.Output('checkbox_confirm', 'value'),
#     [dash.dependencies.Input('Solver_dropdown', 'value'),
#     dash.dependencies.Input('output_bargraph', 'clickData')]
# )
# def check_or_uncheck_checkbox(solver_name, clickData):
#     '''
#     param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
#     param: solver_name (str) - name of the selected solver from the dropdown menu
#     return: checkbox_confirm value (str) - status of mentor/solver pairing
#     '''
#     df = pd.read_excel('mit_solve_confirmed_matches.xlsx')
#     partners_list = df['PARTNER'].tolist()
#     solvers_list = df['SOLVER'].tolist()

#     # check to make sure there is a partner selected
#     if clickData == None:
#             return 'You need to select a partner'

#     # iterate through list of solvers to find currently selected solver (solver_name)
#     for i in range(len(solvers_list)):
#         if solvers_list[i] == solver_name:
#             # if the solver name is found check if its partner is the currently selected partner
#             if partners_list[i] == clickData['points'][0]['label']:
#                 # This is  a match 
#                 return 'Confirm'

#     # If we get here this is not a match
#     return 'Denied'


# # Callback that adds and deletes matches to the 'mit_solve_confirmed_matches.xlsx'
# @app.callback(
#     dash.dependencies.Output('hidden-div', 'children'),
#     [dash.dependencies.Input('checkbox_confirm', 'value')],
#     [dash.dependencies.State('Solver_dropdown', 'value'),
#     dash.dependencies.State('output_bargraph', 'clickData')]
#     )
# def add_confirmed_match(checkbox, solver_name, clickData):
#     '''
#     param: checkbox (str) - defines whether there is a match between partner and solver
#     param: clickData (Plotly Dash Object) - data that is collected from clicking on graph
#     param: solver_name (str) - name of the selected solver from the dropdown menu
#     return: children (str) - irrelavent output, will never be printed out and is used to 
#     comply with needing an Output for every callback
#     '''
#     # Check if we are adding a match
#     if checkbox == 'Confirm':
#         if clickData == None:
#             return 'You need to select a partner'
#         else:
#             df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
#             partners_list = df['PARTNER'].tolist()
#             solvers_list = df['SOLVER'].tolist()

#             # checks if already a match
#             for i in range(len(solvers_list)):
#                 if solvers_list[i] == solver_name:
#                     if partners_list[i] == clickData['points'][0]['label']:
#                         # This is already a match 
#                         return None

#             file = 'mit_solve_confirmed_matches.xlsx'
#             wb = openpyxl.load_workbook(filename=file)
#             ws = wb.get_sheet_by_name('Sheet1')

#             # count number of matches 
#             # start the count at 1 to account for this match not being added to the sheet yet
#             matches_count_for_partner = 1
#             for i in range(len(partners_list)):
#                 if partners_list[i] == clickData['points'][0]['label']:
#                     matches_count_for_partner += 1
                    
#             # insert the partner and solver names, as well as datetime and number of matches
#             time_right_now = datetime.datetime.now()
#             ws['A' + str(COUNT_OF_MATCHES + 2)] = str(clickData['points'][0]['label'])
#             ws['B' + str(COUNT_OF_MATCHES + 2)] = str(solver_name)
#             ws['C' + str(COUNT_OF_MATCHES + 2)] = str(time_right_now)
#             ws['D' + str(COUNT_OF_MATCHES + 2)] = str(matches_count_for_partner)
#             # Save the workbook
#             wb.save(file)
#             #increment amount of total matches
#             increment_count_of_matches()
#             return ''

#     # Check if we are removing a match
#     if checkbox == 'Denied':
#         if clickData == None:
#             return 'No partner selected'
#         else:
#             df = pd.read_excel('mit_solve_confirmed_matches.xlsx') 
#             partners_list = df['PARTNER'].tolist()
#             solvers_list = df['SOLVER'].tolist()

#             # checks if already a match
#             for i in range(len(solvers_list)):
#                 if solvers_list[i] == solver_name:
#                     if partners_list[i] == clickData['points'][0]['label']:
#                         # This match needs to be deleted

#                         file = 'mit_solve_confirmed_matches.xlsx'
#                         wb = openpyxl.load_workbook(filename=file)
#                         # Select the right sheet
#                         ws = wb.get_sheet_by_name('Sheet1')
#                         # insert the partner and solver name, datetime, and number of matches
#                         ws['A' + str(i + 2)] = str('')
#                         ws['B' + str(i + 2)] = str('')
#                         ws['C' + str(i + 2)] = str('')
#                         ws['D' + str(i + 2)] = str('')
#                         # Save the workbook
#                         wb.save(file)
#                         # NEED TO DECREMENT NUMBER OF MATCHES HERE
#                         # LOGIC MAY NEED WORK TOO, NOT SURE IF JUST DECREMENTING IS THE RIGHT MOVE






# # This callback will create a new bar chart with the data from the uploaded excel
# # files instead of the preloaded old excel files
# @app.callback(
#     dash.dependencies.Output('Solver_dropdown', 'value'),
#     [dash.dependencies.Input('upload-data', 'contents')],
# )
# def point_graph_to_uploaded_files(contents):
#     '''
#     param: contents - all of the files uploaded
#     return: solver_name (str) - name of the selected solver from the dropdown menu
#     '''
#     try:
#         # create new df from uploaded file
#         xls_file_total_score = pd.ExcelFile('MIT_SOLVE_downloadable_excel_files/total_score_from_upload.xlsx')
#         uploaded_df_total_score = xls_file_total_score.parse('Sheet1')
#         # Create new graph with uploaded data instead of hardcoded
#         new_solvers = list(uploaded_df_total_score.columns[1:])
#         # Returns a different solver its obvious whether this worked or not
#         return new_solvers[8]
#     except:
#         return Solvers[0]




if __name__ == '__main__':
    app.run_server(debug=True)